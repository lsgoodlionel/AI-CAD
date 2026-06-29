"""
AI 审图报告生成器

- generate_annotated_pdf: 在原图纸 PDF 上标注问题坐标，返回带注释的 PDF 字节
- generate_excel_report:  生成按严重程度分 Sheet 的 Excel 清单，返回 xlsx 字节
"""
import io
import json
from datetime import datetime
from typing import Any

import fitz  # pymupdf
import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
)

# ── 严重程度配置 ───────────────────────────────────────────────

_SEVERITY_COLOR = {
    "critical": (0.9, 0.1, 0.1),   # 红
    "major":    (0.95, 0.55, 0.0),  # 橙
    "minor":    (0.0, 0.5, 0.9),    # 蓝
    "info":     (0.3, 0.7, 0.3),    # 绿
}

_SEVERITY_LABEL = {
    "critical": "严重（强条）",
    "major":    "重大",
    "minor":    "一般",
    "info":     "建议",
}

_SEVERITY_XLSX_COLOR = {
    "critical": "FFCCCC",
    "major":    "FFE5CC",
    "minor":    "CCE5FF",
    "info":     "CCFFCC",
}


# ── PDF 批注生成 ───────────────────────────────────────────────

def generate_annotated_pdf(
    original_pdf_bytes: bytes,
    issues: list[dict[str, Any]],
) -> bytes:
    """
    在原 PDF 每页上标注 AI 审查问题。
    issues 字段: engine, severity, category, description, location_x (0-1), location_y (0-1)
    """
    doc = fitz.open(stream=original_pdf_bytes, filetype="pdf")

    # 按页码分组（无坐标的放到第 0 页末尾）
    page_count = len(doc)
    issues_with_coords = [i for i in issues if i.get("location_x") is not None and i.get("location_y") is not None]
    issues_no_coords   = [i for i in issues if i.get("location_x") is None or i.get("location_y") is None]

    for page_idx in range(page_count):
        page = doc[page_idx]
        w, h = page.rect.width, page.rect.height

        # 标注有坐标的问题
        for seq, issue in enumerate(issues_with_coords, start=1):
            px = float(issue["location_x"]) * w
            py = float(issue["location_y"]) * h
            color = _SEVERITY_COLOR.get(issue.get("severity", "info"), (0.3, 0.7, 0.3))

            # 圆形标注
            radius = 12
            circle_rect = fitz.Rect(px - radius, py - radius, px + radius, py + radius)
            annot = page.add_circle_annot(circle_rect)
            annot.set_colors(stroke=color, fill=(*color, 0.3))
            annot.set_info(title=f"[{issue.get('severity','').upper()}] {issue.get('category','')}", content=issue.get("description", ""))
            annot.update(opacity=0.8)

            # 序号标签
            page.draw_circle(fitz.Point(px, py), radius, color=color, fill=color, width=1.5)
            page.insert_text(
                fitz.Point(px - (6 if seq < 10 else 10), py + 4),
                str(seq),
                fontsize=10,
                color=(1, 1, 1),
            )

    # 无坐标问题附到最后一页底部列表
    if issues_no_coords and page_count > 0:
        last_page = doc[-1]
        _append_issue_list(last_page, issues_no_coords)

    # 加图例到第一页右下角
    if issues and page_count > 0:
        _draw_legend(doc[0])

    buf = io.BytesIO()
    doc.save(buf, garbage=4, deflate=True)
    doc.close()
    return buf.getvalue()


def _draw_legend(page: fitz.Page) -> None:
    """在页面右下角绘制图例"""
    w, h = page.rect.width, page.rect.height
    x0, y0 = w - 140, h - 100
    page.draw_rect(fitz.Rect(x0 - 5, y0 - 15, w - 5, h - 5), color=(0, 0, 0), fill=(1, 1, 1), width=0.5)
    page.insert_text(fitz.Point(x0, y0), "AI 审图图例", fontsize=8, color=(0, 0, 0))
    for i, (sev, color) in enumerate(_SEVERITY_COLOR.items()):
        row_y = y0 + 14 + i * 16
        page.draw_circle(fitz.Point(x0 + 6, row_y - 4), 5, color=color, fill=color)
        page.insert_text(fitz.Point(x0 + 15, row_y), _SEVERITY_LABEL[sev], fontsize=7, color=(0, 0, 0))


def _append_issue_list(page: fitz.Page, issues: list[dict]) -> None:
    """在页面底部追加无坐标问题列表"""
    _, h = page.rect.width, page.rect.height
    y = h - 20 - len(issues) * 14
    if y < 50:
        y = 50
    page.insert_text(fitz.Point(20, y - 14), "以下问题无坐标信息：", fontsize=8, color=(0.4, 0.4, 0.4))
    for issue in issues:
        color = _SEVERITY_COLOR.get(issue.get("severity", "info"), (0, 0, 0))
        label = f"[{issue.get('severity','').upper()}] {issue.get('description','')[:80]}"
        page.insert_text(fitz.Point(20, y), label, fontsize=7, color=color)
        y += 13


# ── Excel 清单生成 ─────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_COLUMNS = [
    ("序号", 6),
    ("严重程度", 12),
    ("引擎", 10),
    ("分类", 16),
    ("问题描述", 45),
    ("规范条文引用", 20),
    ("整改建议", 35),
    ("处理状态", 10),
]


def generate_excel_report(
    issues: list[dict[str, Any]],
    drawing_no: str,
    discipline: str,
    report_date: datetime | None = None,
) -> bytes:
    """生成按严重程度分 Sheet 的 Excel 问题清单"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    report_date = report_date or datetime.now()
    date_str = report_date.strftime("%Y-%m-%d")

    # 汇总 Sheet
    _build_summary_sheet(wb, issues, drawing_no, discipline, date_str)

    # 按严重程度分 Sheet
    severity_order = ["critical", "major", "minor", "info"]
    for sev in severity_order:
        sev_issues = [i for i in issues if i.get("severity") == sev]
        if sev_issues:
            _build_issue_sheet(wb, sev_issues, sev, drawing_no, date_str)

    # 会审问题单 Sheet（仅当存在会审审查引擎问题时）
    review_issues = [i for i in issues if i.get("engine") == "review"]
    if review_issues:
        _build_review_sheet(wb, review_issues)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(
    wb: openpyxl.Workbook,
    issues: list[dict],
    drawing_no: str,
    discipline: str,
    date_str: str,
) -> None:
    ws = wb.create_sheet("汇总")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    header_fill = PatternFill("solid", fgColor="2F5496")
    ws["A1"] = "AI 审图报告汇总"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:B1")

    meta = [
        ("图纸编号", drawing_no),
        ("专业", discipline),
        ("生成日期", date_str),
        ("问题总数", len(issues)),
    ]
    for row_i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=row_i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row_i, column=2, value=v)

    # 各严重程度统计
    ws.cell(row=7, column=1, value="严重程度").font = Font(bold=True)
    ws.cell(row=7, column=2, value="数量").font = Font(bold=True)
    for row_i, sev in enumerate(["critical", "major", "minor", "info"], start=8):
        count = sum(1 for i in issues if i.get("severity") == sev)
        ws.cell(row=row_i, column=1, value=_SEVERITY_LABEL.get(sev, sev))
        cell = ws.cell(row=row_i, column=2, value=count)
        if count > 0:
            cell.fill = PatternFill("solid", fgColor=_SEVERITY_XLSX_COLOR[sev])


def _build_issue_sheet(
    wb: openpyxl.Workbook,
    issues: list[dict],
    severity: str,
    drawing_no: str,
    date_str: str,
) -> None:
    sheet_name = _SEVERITY_LABEL.get(severity, severity)[:15]
    ws = wb.create_sheet(sheet_name)

    # 设置列宽
    for col_idx, (_, width) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 标题行
    title_fill = PatternFill("solid", fgColor=_SEVERITY_XLSX_COLOR.get(severity, "EEEEEE"))
    ws.row_dimensions[1].height = 30
    for col_idx, (col_name, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL if severity in ("critical", "major") else PatternFill("solid", fgColor="2F5496")
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # 数据行
    row_fill = PatternFill("solid", fgColor=_SEVERITY_XLSX_COLOR.get(severity, "FFFFFF"))
    status_labels = {"open": "待处理", "acknowledged": "已知晓", "closed": "已关闭", "waived": "已豁免"}

    for row_idx, issue in enumerate(issues, start=2):
        row_data = [
            row_idx - 1,
            _SEVERITY_LABEL.get(severity, severity),
            issue.get("engine", ""),
            issue.get("category", ""),
            issue.get("description", ""),
            issue.get("regulation_ref", ""),
            issue.get("suggestion", ""),
            status_labels.get(issue.get("status", "open"), issue.get("status", "")),
        ]
        ws.row_dimensions[row_idx].height = 40
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER
            if col_idx > 1:
                cell.fill = row_fill


# ── 会审问题单（会审审查引擎 engine=='review'）─────────────────────

_REVIEW_COLUMNS: list[tuple[str, int]] = [
    ("序号", 6), ("专业", 16), ("风险等级", 10), ("问题归类", 18),
    ("标准问题", 60), ("接口复核", 20), ("证据缺口", 30),
    # ── V2 列（向后兼容：旧 issue 无 V2 字段时留空）──
    ("场景", 12), ("对象", 18), ("主问题", 60), ("补充问题", 60),
]


def _as_list(value: Any) -> list:
    """JSONB 字段经 DB 驱动可能为 list 或 JSON 文本，统一为 list。"""
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, list) else [parsed]
        except (json.JSONDecodeError, ValueError):
            return [value] if value else []
    return [value]


def _as_dict(value: Any) -> dict:
    """JSONB 字段经 DB 驱动可能为 dict 或 JSON 文本，统一为 dict；解析失败返回空 dict。"""
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, ValueError):
            return {}
    return {}


def _build_review_sheet(wb: openpyxl.Workbook, issues: list[dict]) -> None:
    """汇总会审审查问题为「会审问题单」工作表。"""
    ws = wb.create_sheet("会审问题单")
    for col_idx, (name, width) in enumerate(_REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="0E7490")
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 28

    for row_idx, issue in enumerate(issues, start=2):
        iface = "、".join(
            x for x in [issue.get("interface_primary", "")] + _as_list(issue.get("interface_related"))
            if x
        )
        question_pack = _as_dict(issue.get("question_pack"))
        row_data = [
            row_idx - 1,
            issue.get("discipline_code", "") or "未分类",
            issue.get("risk_level", "") or "—",
            "/".join(str(x) for x in _as_list(issue.get("issue_class"))),
            issue.get("standard_question", "") or issue.get("description", ""),
            iface,
            "；".join(str(x) for x in _as_list(issue.get("evidence_gap"))),
            # ── V2 列（缺失字段留空，不报错）──
            issue.get("scenario", "") or "",
            issue.get("object_name", "") or "",
            question_pack.get("主问题", "") or "",
            question_pack.get("补充问题", "") or "",
        ]
        ws.row_dimensions[row_idx].height = 42
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER
